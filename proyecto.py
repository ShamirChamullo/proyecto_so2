import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import streamlit as st
from pathlib import Path
import tempfile

def process_files(files, start_col, end_col, start_row, temp_dir):
    start_col_index = column_index_from_string(start_col)
    end_col_index = column_index_from_string(end_col)
    all_data = []

    for uploaded_file in files:
        file_path = uploaded_file.name
        match = re.search(r'AvanceVentasINTI\.(\d{4})\.(\d{2})\.(\d{2})\.', file_path)
        if match:
            year, month, day = match.groups()
        else:
            year, month, day = "", "", ""

        wb = load_workbook(filename=uploaded_file, read_only=True)
        sheet = wb['ITEM_O']

        data = []
        for row in sheet.iter_rows(min_row=start_row, min_col=start_col_index, max_col=end_col_index):
            data.append([cell.value for cell in row])

        df = pd.DataFrame(data)
        df['ANIO'] = year
        df['MES'] = month
        df['DIA'] = day

        all_data.append(df)

    final_df = pd.concat(all_data, ignore_index=True)

    column_names = [get_column_letter(i) for i in range(start_col_index, end_col_index + 1)]
    column_names.extend(['ANIO', 'MES', 'DIA'])
    final_df.columns = column_names

    output_path = os.path.join(temp_dir, 'Out.xlsx')
    final_df.to_excel(output_path, index=False)

    generate_charts(final_df, temp_dir)

    return final_df, output_path

def generate_charts(df, temp_dir):
    chart_folder = os.path.join(temp_dir, 'charts')
    if not os.path.exists(chart_folder):
        os.makedirs(chart_folder)

    for column in df.columns:
        if df[column].dtype in ['int64', 'float64']:
            plt.figure()
            df[column].hist()
            plt.title(f'Histograma de {column}')
            hist_path = os.path.join(chart_folder, f'{column}_hist.png')
            plt.savefig(hist_path)
            plt.close()

        if df[column].dtype == 'object' or df[column].dtype.name == 'category':
            plt.figure()
            df[column].value_counts().plot.pie(autopct='%1.1f%%')
            plt.title(f'Torta de {column}')
            pie_path = os.path.join(chart_folder, f'{column}_pie.png')
            plt.savefig(pie_path)
            plt.close()

# Interfaz de usuario con Streamlit
st.title("Proceso ETL")

uploaded_files = st.file_uploader("Seleccione archivos de la carpeta deseada:", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    start_column = st.text_input("Columna inicial (ej. A):").upper()
    end_column = st.text_input("Columna final (ej. P):").upper()
    start_row = st.number_input("Fila inicial:", min_value=1, step=1, value=1)

    if st.button("Procesar archivos"):
        if start_column and end_column and start_row:
            with st.spinner("Procesando archivos..."):
                try:
                    # Create a temporary directory
                    with tempfile.TemporaryDirectory() as temp_dir:
                        final_df, output_file_path = process_files(uploaded_files, start_column, end_column, start_row, temp_dir)
                        
                        # Provide download links
                        st.success("Proceso completado.")
                        st.dataframe(final_df.head())
                        
                        # Provide link to download the excel file
                        with open(output_file_path, 'rb') as f:
                            st.download_button(
                                label="Descargar archivo Excel",
                                data=f,
                                file_name='Out.xlsx'
                            )

                        # Provide links to download the charts
                        chart_folder = os.path.join(temp_dir, 'charts')
                        for image_file in os.listdir(chart_folder):
                            image_path = os.path.join(chart_folder, image_file)
                            with open(image_path, 'rb') as img:
                                st.download_button(
                                    label=f"Descargar {image_file}",
                                    data=img,
                                    file_name=image_file
                                )
                        
                except Exception as e:
                    st.error(f"Error: {str(e)}")
        else:
            st.warning("Por favor, complete todos los campos.")
else:
    st.info("Por favor, seleccione archivos de la carpeta que desea procesar.")
